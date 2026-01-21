from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import shutil
import warnings
from dataclasses import asdict

pd.options.display.float_format = '{:12.3e}'.format

from src.utils import formatter

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.precision', 3)

DIRECTORY_NAME = "data"
FILE_NAME = "Leiterseildaten.csv"
EXCEL_INPUT_FILE = "Eingaben.xlsx"

def get_project_root() -> Path:
    return Path(__file__).parent.parent.parent

def load_csv_to_df() -> pd.DataFrame:
    file = Path(get_project_root(), DIRECTORY_NAME, FILE_NAME)
    if file:
        raw_data = pd.read_csv(file, header=0, delimiter=";", na_filter=False)
        df = pd.DataFrame(raw_data)
    else:
        print(f"Datei {FILE_NAME} im Dateiordner {DIRECTORY_NAME} konnte nicht gefunden oder geladen werden. Prüfe "
              f"den Pfad {file}!")
        df = pd.DataFrame()
    return df

def convert_df_to_dict(df: pd.DataFrame) -> list[dict]:
    dictionary = df.to_dict(orient='records')
    return dictionary

def load_excel_to_df(file_path: str | Path = None) -> pd.DataFrame:
    """
    Lädt eine Excel-Datei und gibt sie als DataFrame zurück.
    Die Excel-Datei hat eine vertikale Struktur:
    - Spalte 0: Parameternamen
    - Spalte 1: Werte

    Args:
        file_path: Pfad zur Excel-Datei. Wenn None, wird die Standard-Eingabe-Datei verwendet.

    Returns:
        DataFrame mit den geladenen Daten (keine Header-Zeile)
    """
    if file_path is None:
        file_path = Path(get_project_root(), DIRECTORY_NAME, EXCEL_INPUT_FILE)
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"Datei {file_path} konnte nicht gefunden werden!")
        return pd.DataFrame()

    try:
        # Verwende Context Manager für automatisches Schließen
        with pd.ExcelFile(file_path, engine='openpyxl') as xlsx:
            raw_data = pd.read_excel(xlsx, header=None, na_filter=False)
            df = pd.DataFrame(raw_data)
        return df

    except FileNotFoundError as e:
        print(f"Excel-Datei nicht gefunden: {e}")
        return pd.DataFrame()
    except PermissionError as e:
        print(f"Keine Leseberechtigung für Excel-Datei: {e}")
        return pd.DataFrame()
    except ValueError as e:
        print(f"Ungültiges Excel-Format oder fehlerhafte Daten: {e}")
        return pd.DataFrame()
    except Exception as e:
        print(f"Unerwarteter Fehler beim Laden der Excel-Datei {file_path}: {e}")
        return pd.DataFrame()

def convert_excel_to_input_dict(df: pd.DataFrame) -> tuple[dict, list[str], list[str]]:
    """
    Konvertiert einen DataFrame aus der Excel-Eingabedatei in ein Dictionary,
    das direkt zum Setzen der Taipy GUI Widgets verwendet werden kann.

    Die Excel-Datei hat eine vertikale Struktur:
    - Spalte 0: Parameternamen (z.B. "Art der Leiterseilbefestigung")
    - Spalte 1: Werte (z.B. "Abgespannt")

    Args:
        df: DataFrame mit den Eingabedaten (vertikales Layout)

    Returns:
        Tuple mit (input_dict, loaded_fields, skipped_fields)
        - input_dict: Dictionary mit den Eingabewerten
        - loaded_fields: Liste der erfolgreich geladenen Felder
        - skipped_fields: Liste der übersprungenen Felder (leer oder nicht gefunden)
    """
    if df.empty:
        return {}, [], []

    # Erstelle ein Dictionary aus den Zeilen: Spalte 0 = Key, Spalte 1 = Value
    excel_data = {}
    for idx, row in df.iterrows():
        if pd.notna(row[0]) and row[0] != '':
            key = str(row[0]).strip()
            value = row[1] if len(row) > 1 and pd.notna(row[1]) and row[1] != '' else None
            if value is not None:
                excel_data[key] = value

    # Erstelle ein Dictionary mit allen relevanten Feldern
    input_dict = {}
    loaded_fields = []
    skipped_fields = []

    # Mapping der Excel-Zeilennamen zu den State-Variablennamen
    field_mapping = {
        'Art der Leiterseilbefestigung': 'leiterseilbefestigung_selected',
        'Schlaufe in Spannfeldmitte': 'schlaufe_in_spannfeldmitte_selected',
        'Höhenunterschied der Befestigungspunkte mehr als 25%': 'hoehenunterschied_befestigungspunkte_selected',
        'Schlaufebene bei Schlaufen in Spannfeldmitte': 'schlaufenebene_parallel_senkrecht_selected',
        'ϑ_l Niedrigste Temperatur': 'temperatur_niedrig_selected',
        'ϑ_h Höchste Temperatur': 'temperatur_hoch_selected',
        'I\'\'k Anfangs-Kurzschlusswechselstrom beim dreipoligen Kurzschluss (Effektivwert)': 'standardkurzschlussstroeme_selected',
        'κ Sossfaktor': 'kappa',
        'Tk Kurzschlussdauer': 't_k',
        'f Frequenz des Netzes': 'frequenz_des_netzes_selected',
        'Leiterseiltyp': 'leiterseiltyp_selected',
        'n Anzahl der Teilleiter eines Hauptleiters': 'teilleiter_selected',
        'm_c Summe konzentrischer Massen im Spannfeld': 'm_c',
        'l Mittenabstand der Stützpunkte': 'l',
        'l_i Länge einer Abspann-Isolatorkette': 'l_i',
        'l_h_f Länge einer Klemme u. Formfaktor': 'l_h_f',
        'a Leitermittenabstand': 'a',
        'a_s wirksamer Abstand zwischen Teilleitern': 'a_s',
        'Fst-20 statische Seilzugkraft in einem Hauptleiter': 'F_st_20',
        'Fst80 statische Seilzugkraft in einem Hauptleiter': 'F_st_80',
        'S resultierender Federkoeffizient beider Stützpunkte im Spannfeld': 'federkoeffizient_selected',
        'Abstand Phasenabstandshalter 1': 'l_s_1',
        'Abstand Phasenabstandshalter 2': 'l_s_2',
        'Abstand Phasenabstandshalter 3': 'l_s_3',
        'Abstand Phasenabstandshalter 4': 'l_s_4',
        'Abstand Phasenabstandshalter 5': 'l_s_5',
        'Abstand Phasenabstandshalter 6': 'l_s_6',
        'Abstand Phasenabstandshalter 7': 'l_s_7',
        'Abstand Phasenabstandshalter 8': 'l_s_8',
        'Abstand Phasenabstandshalter 9': 'l_s_9',
        'Abstand Phasenabstandshalter 10': 'l_s_10',
    }

    # Konvertiere alle vorhandenen Felder
    for excel_label, state_var in field_mapping.items():
        if excel_label in excel_data:
            value = excel_data[excel_label]
            input_dict[state_var] = value
            loaded_fields.append(excel_label)
        else:
            skipped_fields.append(excel_label)

    return input_dict, loaded_fields, skipped_fields

def _convert_value(value):
    """
    Konvertiert einen Wert in den passenden Typ für Excel-Export.

    Args:
        value: Zu konvertierender Wert

    Returns:
        Konvertierter Wert (str, int, float oder '')
    """
    # Prüfe auf leere Werte
    if value is None or value in ('0', 0):
        return ''

    # Wenn bereits eine Zahl, direkt zurückgeben
    if isinstance(value, (int, float)):
        return value

    # String-Konvertierung versuchen
    if isinstance(value, str):
        try:
            # Float wenn Dezimaltrennzeichen vorhanden
            if '.' in value or ',' in value:
                return float(value.replace(',', '.'))
            # Ansonsten Integer
            return int(value)
        except (ValueError, AttributeError):
            # Bei Fehler String beibehalten
            return value

    return value


def export_input_dict_to_excel(input_dict: dict, template_path: str | Path, output_path: str | Path) -> bool:
    """
    Exportiert Dictionary in Excel-Datei basierend auf Vorlage.

    Args:
        input_dict: Dictionary mit State-Variablennamen als Keys und Werten
        template_path: Pfad zur Vorlagen-Excel-Datei
        output_path: Pfad zur Ausgabe-Excel-Datei

    Returns:
        True bei Erfolg, False bei Fehler
    """
    warnings.filterwarnings("ignore", category=ResourceWarning, message=".*unclosed file.*")

    template_path = Path(template_path)
    if not template_path.exists():
        print(f"Vorlage nicht gefunden: {template_path}")
        return False

    reverse_field_mapping = {
        'leiterseilbefestigung_selected': 'Art der Leiterseilbefestigung',
        'schlaufe_in_spannfeldmitte_selected': 'Schlaufe in Spannfeldmitte',
        'hoehenunterschied_befestigungspunkte_selected': 'Höhenunterschied der Befestigungspunkte mehr als 25%',
        'schlaufenebene_parallel_senkrecht_selected': 'Schlaufebene bei Schlaufen in Spannfeldmitte',
        'temperatur_niedrig_selected': 'ϑ_l Niedrigste Temperatur',
        'temperatur_hoch_selected': 'ϑ_h Höchste Temperatur',
        'standardkurzschlussstroeme_selected': 'I\'\'k Anfangs-Kurzschlusswechselstrom beim dreipoligen Kurzschluss (Effektivwert)',
        'kappa': 'κ Sossfaktor',
        't_k': 'Tk Kurzschlussdauer',
        'frequenz_des_netzes_selected': 'f Frequenz des Netzes',
        'leiterseiltyp_selected': 'Leiterseiltyp',
        'teilleiter_selected': 'n Anzahl der Teilleiter eines Hauptleiters',
        'm_c': 'm_c Summe konzentrischer Massen im Spannfeld',
        'l': 'l Mittenabstand der Stützpunkte',
        'l_i': 'l_i Länge einer Abspann-Isolatorkette',
        'l_h_f': 'l_h_f Länge einer Klemme u. Formfaktor',
        'a': 'a Leitermittenabstand',
        'a_s': 'a_s wirksamer Abstand zwischen Teilleitern',
        'F_st_20': 'Fst-20 statische Seilzugkraft in einem Hauptleiter',
        'F_st_80': 'Fst80 statische Seilzugkraft in einem Hauptleiter',
        'federkoeffizient_selected': 'S resultierender Federkoeffizient beider Stützpunkte im Spannfeld',
        'l_s_1': 'Abstand Phasenabstandshalter 1',
        'l_s_2': 'Abstand Phasenabstandshalter 2',
        'l_s_3': 'Abstand Phasenabstandshalter 3',
        'l_s_4': 'Abstand Phasenabstandshalter 4',
        'l_s_5': 'Abstand Phasenabstandshalter 5',
        'l_s_6': 'Abstand Phasenabstandshalter 6',
        'l_s_7': 'Abstand Phasenabstandshalter 7',
        'l_s_8': 'Abstand Phasenabstandshalter 8',
        'l_s_9': 'Abstand Phasenabstandshalter 9',
        'l_s_10': 'Abstand Phasenabstandshalter 10',
    }

    # Kopiere die Vorlage zur Ausgabedatei (um Formatierung zu erhalten)
    shutil.copy(template_path, output_path)

    # Öffne die kopierte Datei mit openpyxl (behält Formatierung)
    wb = None
    try:
        wb = load_workbook(output_path, keep_vba=False, data_only=False)
        ws = wb.active

        # Durchlaufe alle Zeilen in der Excel-Datei
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            # Spalte A (Index 0) enthält die Labels
            if row[0].value:
                excel_label = str(row[0].value).strip()

                # Suche nach dem passenden State-Variablennamen
                for state_var, mapped_label in reverse_field_mapping.items():
                    if excel_label == mapped_label:
                        # Hole den Wert aus dem input_dict
                        value = input_dict.get(state_var, '')

                        # Konvertiere den Wert in den richtigen Typ
                        if value is None or value == '0' or value == 0:
                            value = ''
                        elif isinstance(value, str):
                            # Versuche String in Zahl zu konvertieren wenn möglich
                            try:
                                # Versuche zuerst Float
                                if '.' in value or ',' in value:
                                    value = float(value.replace(',', '.'))
                                else:
                                    # Versuche Int
                                    value = int(value)
                            except (ValueError, AttributeError):
                                # Wenn nicht konvertierbar, behalte String
                                pass

                        # Setze den Wert in Spalte B (Index 1)
                        row[1].value = value
                        break

        # Speichere die Datei (Formatierung bleibt erhalten)
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Fehler beim Exportieren der Excel-Datei: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # Stelle sicher, dass die Datei geschlossen wird, auch bei Fehler
        if wb is not None:
            wb.close()

def create_df_from_calc_results(calc_result: dict, temp_low: str, temp_high: str) -> pd.DataFrame:
    """
    Erstellt einen formatierten DataFrame aus den Berechnungsergebnissen.

    Args:
        calc_result: Dictionary mit 'F_st_20' und 'F_st_80' als Keys
        temp_low: String of the selected temperature
        temp_high: String of the selected temperature
    Returns:
        DataFrame mit formatierten Ergebnissen
    """
    temp_low_with_unit = str(temp_low) + " " + "°C"
    temp_high_with_unit = str(temp_high) + " " + "°C"

    # Hole beide Ergebnisse
    result_20 = calc_result.get('F_st_20')
    result_80 = calc_result.get('F_st_80')

    if not result_20 or not result_80:
        return pd.DataFrame({'Hinweis': ['Keine Berechnungsergebnisse verfügbar']})

    # Konvertiere zu Dictionaries
    dict_20 = asdict(result_20)
    dict_80 = asdict(result_80)

    # Erstelle Liste für DataFrame-Rows
    data = []

    # Iteriere durch alle Felder
    for param_name in dict_20.keys():
        val_20 = dict_20[param_name]
        val_80 = dict_80[param_name]

        # Nur nicht-None Werte hinzufügen
        if val_20 is not None and val_80 is not None:
            # WICHTIG: Konvertiere Sympy-Objekte zu nativen Python-Typen
            # Sympy.Float ist nicht JSON-serialisierbar für Taipy
            if hasattr(val_20, '__float__'):
                val_20 = float(val_20)
            if hasattr(val_80, '__float__'):
                val_80 = float(val_80)

            # Formatiere mit der Formatter-Funktion
            val_20_formatted = formatter.format_value_smart(val_20)
            val_80_formatted = formatter.format_value_smart(val_80)

            data.append({
                'Parameter': param_name,
                temp_low_with_unit: val_20_formatted,
                temp_high_with_unit: val_80_formatted
                #'-20°C': val_20,
                #'80°C': val_80
            })
    # Erstelle DataFrame
    df = pd.DataFrame(data)

    # Unindent this part and the part above if for numerical export needed
    """
        if not df.empty:
            df['Parameter'] = df['Parameter'].astype(str)
            df['-20°C'] = pd.to_numeric(df['-20°C'], errors='coerce')
            df['80°C'] = pd.to_numeric(df['80°C'], errors='coerce')
    """

    return df




if __name__ == "__main__":
    # Liefert alle Werte einer Spalte zurück
    #print(load_csv_to_df()["Bezeichnung"])

    # Liefert die Bezeichnung eines speziellen Indexes zurück
    #print(load_csv_to_df().at[2, "Bezeichnung"])

    # Liefert den Wert einer Spalte basierend auf dem Index zurück
    #print(load_csv_to_df().iloc[2]["Querschnitt eines Teilleiters"])

    #print(load_csv_to_df())

    print(convert_df_to_dict(load_csv_to_df()))
    list_dict_leiterseile = convert_df_to_dict(load_csv_to_df())
    print(list_dict_leiterseile[0])
    dict_leiterseile = list_dict_leiterseile[0]

    #print(list(dict_leiterseile.keys()))
    #print(list(dict_leiterseile.values()))
    print(dict_leiterseile.get("Bezeichnung"))
    #print(dict_leiterseile["Bezeichnung"])
